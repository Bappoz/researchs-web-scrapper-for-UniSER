"""
🧪 TESTE MANUAL - EXCEL CONSOLIDADO
Cria um arquivo Excel consolidado para teste manual
"""

import sys
sys.path.append('.')
from src.database.excel_consolidado import ConsolidatedExcelExporter
from src.database.mongodb import research_db
import os

def create_test_excel():
    """Criar Excel de teste para abertura manual"""
    
    print("🧪 Criando Excel consolidado para teste manual...")
    
    try:
        # Obter dados do MongoDB
        research_data = research_db.get_all_keyword_filtered_research()
        
        if not research_data:
            print("❌ Nenhum dado encontrado no MongoDB")
            return None
        
        print(f"📊 Encontrados dados de {len(research_data)} pesquisas")
        
        # Criar exporter
        exporter = ConsolidatedExcelExporter()
        
        # Gerar arquivo Excel
        filename = exporter.export_consolidated_excel(research_data)
        
        if filename:
            filepath = os.path.join(exporter.exports_dir, filename)
            print(f"✅ Arquivo criado: {filepath}")
            print(f"📁 Tamanho: {os.path.getsize(filepath)} bytes")
            
            # Verificar se é válido
            from openpyxl import load_workbook
            wb = load_workbook(filepath)
            print(f"📋 Abas encontradas: {wb.sheetnames}")
            
            # Mostrar estatísticas das abas
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = ws.max_row
                cols = ws.max_column
                print(f"  📄 {sheet_name}: {rows} linhas, {cols} colunas")
            
            print(f"\n🎯 TESTE MANUAL:")
            print(f"   Navegue até: {filepath}")
            print(f"   Tente abrir o arquivo Excel diretamente")
            print(f"   Se não conseguir, tente copiar para a área de trabalho")
            
            return filepath
        else:
            print("❌ Falha na criação do arquivo")
            return None
            
    except Exception as e:
        print(f"❌ Erro: {e}")
        import traceback
        traceback.print_exc()
        return None

if __name__ == "__main__":
    create_test_excel()