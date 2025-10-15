"""
📊 EXCEL CONSOLIDADO - MONGODB
Exportador Excel consolidado usando dados do MongoDB
"""

import os
from datetime import datetime
from typing import List, Dict, Any
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from .mongodb import research_db

class ConsolidatedExcelExporter:
    """Exportador Excel consolidado do MongoDB"""
    
    def __init__(self):
        self.exports_dir = os.path.join(os.getcwd(), "exports")
        if not os.path.exists(self.exports_dir):
            os.makedirs(self.exports_dir)
        
        # Keywords para identificação
        self.KEYWORDS = [
            # Português
            "idoso", "idosos", "idosa", "idosas", "envelhecimento", "envelhecer",
            "terceira idade", "melhor idade", "idade avançada", "pessoa idosa",
            "geriátrico", "geriátrica", "geriatria", "gerontologia", "gerontológico",
            "alzheimer", "demência", "demência senil", "parkinson", "fragilidade",
            "sarcopenia", "osteoporose", "quedas", "institucionalização",
            
            # English
            "elderly", "elder", "elders", "aging", "ageing", "aged", "senior", "seniors",
            "older adult", "older adults", "older people", "geriatric", "geriatrics",
            "gerontology", "gerontological", "alzheimer", "dementia", "parkinson",
            "frailty", "sarcopenia", "osteoporosis", "falls", "institutionalization",
            
            # Español
            "anciano", "ancianos", "anciana", "ancianas", "envejecimiento", "envejecer",
            "tercera edad", "personas mayores", "geriátrico", "geriátrica", "geriatría",
            "gerontología", "alzheimer", "demencia", "parkinson", "fragilidad"
        ]
    
    def export_consolidated_excel(self, research_data: List[Dict] = None, include_stats: bool = True) -> str:
        """Exportar Excel consolidado com todas as pesquisas filtradas"""
        try:
            print("📊 Iniciando exportação consolidada do MongoDB...")
            
            # Buscar dados do MongoDB se não fornecidos
            if research_data is None:
                research_data = research_db.get_all_keyword_filtered_research()
            
            if not research_data:
                print("❌ Nenhum dado encontrado no MongoDB")
                return None
            
            print(f"📚 Processando {len(research_data)} pesquisas...")
            
            # Preparar dados para Excel
            all_publications = []
            researchers_summary = []
            
            for research in research_data:
                researcher_info = research.get("researcher_info", {})
                
                # Buscar publicações em diferentes campos (compatibilidade com estruturas antigas/novas)
                publications = research.get("publications", [])
                if not publications:
                    publications = research.get("data", {}).get("publications", [])
                
                # Adicionar informações do pesquisador ao resumo
                timestamp = research.get("timestamp", "")
                if timestamp:
                    # Se for datetime object, converter para string
                    if hasattr(timestamp, 'strftime'):
                        timestamp_str = timestamp.strftime("%Y-%m-%d")
                    elif isinstance(timestamp, str):
                        timestamp_str = timestamp.split("T")[0] if "T" in timestamp else timestamp
                    else:
                        timestamp_str = str(timestamp)
                else:
                    timestamp_str = "N/A"
                
                researchers_summary.append({
                    "Nome": researcher_info.get("name", "N/A"),
                    "Instituição": researcher_info.get("institution", "N/A"),
                    "H-Index": researcher_info.get("h_index", "N/A"),
                    "i10-Index": researcher_info.get("i10_index", "N/A"),
                    "Total de Citações": researcher_info.get("total_citations", "N/A"),
                    "Plataforma": research.get("platform", "N/A"),
                    "Total de Publicações": research.get("total_publications", 0),
                    "Data da Pesquisa": timestamp_str,
                    "Tempo de Execução (s)": research.get("execution_time", 0)
                })
                
                # Processar publicações
                for pub in publications:
                    # Encontrar keywords na publicação
                    keywords_found = self._find_keywords_in_publication(pub)
                    
                    # Processar timestamp da mesma forma
                    timestamp = research.get("timestamp", "")
                    if timestamp:
                        if hasattr(timestamp, 'strftime'):
                            timestamp_str = timestamp.strftime("%Y-%m-%d")
                        elif isinstance(timestamp, str):
                            timestamp_str = timestamp.split("T")[0] if "T" in timestamp else timestamp
                        else:
                            timestamp_str = str(timestamp)
                    else:
                        timestamp_str = "N/A"
                    
                    # Corrigir autores: se "authors" contém apenas o pesquisador principal,
                    # e "publication" contém mais autores, usar a publication para autores
                    authors_field = pub.get("authors", researcher_info.get("name", "N/A"))
                    publication_field = pub.get("publication", "N/A")
                    
                    # Se authors é só o nome do pesquisador e publication tem mais informação
                    if (authors_field == researcher_info.get("name", "") and 
                        publication_field != "N/A" and 
                        len(publication_field) > len(authors_field)):
                        # Publication contém os autores reais, extrair apenas os autores
                        if " - " in publication_field:
                            # Formato: "Autores - Revista"
                            corrected_authors = publication_field.split(" - ")[0].strip()
                            corrected_publication = publication_field.split(" - ", 1)[1].strip()
                        else:
                            # Se não tem separador, usar publication como autores
                            corrected_authors = publication_field
                            corrected_publication = publication_field
                    else:
                        # Usar os campos como estão
                        corrected_authors = authors_field
                        corrected_publication = publication_field
                    
                    publication_data = {
                        "Pesquisador": researcher_info.get("name", "N/A"),
                        "Instituição": researcher_info.get("institution", "N/A"),
                        "Título": pub.get("title", "N/A"),
                        "Autores": corrected_authors,
                        "Publicação/Revista": corrected_publication,
                        "Ano": pub.get("year", "N/A"),
                        "Citações": pub.get("cited_by", 0),
                        "Tipo": pub.get("type", "N/A"),
                        "Plataforma": pub.get("platform", research.get("platform", "N/A")),
                        "Keywords Encontradas": ", ".join(keywords_found) if keywords_found else "N/A",
                        "Data da Coleta": timestamp_str
                    }
                    
                    all_publications.append(publication_data)
            
            # Criar Excel
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"excel_consolidado_{timestamp}.xlsx"
            filepath = os.path.join(self.exports_dir, filename)
            
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                # Aba 1: Todas as Publicações
                if all_publications:
                    df_publications = pd.DataFrame(all_publications)
                    df_publications.to_excel(writer, sheet_name='Publicações', index=False)
                    
                    # Formatar aba de publicações
                    self._format_publications_sheet(writer.sheets['Publicações'], df_publications)
                
                # Aba 2: Resumo dos Pesquisadores
                if researchers_summary:
                    df_researchers = pd.DataFrame(researchers_summary)
                    df_researchers.to_excel(writer, sheet_name='Pesquisadores', index=False)
                    
                    # Formatar aba de pesquisadores
                    self._format_researchers_sheet(writer.sheets['Pesquisadores'], df_researchers)
                
                # Aba 3: Estatísticas (se solicitado)
                if include_stats:
                    try:
                        stats = research_db.get_research_statistics()
                        self._create_statistics_sheet(writer, stats, len(all_publications), len(researchers_summary))
                    except Exception as stats_error:
                        print(f"⚠️ Erro ao obter estatísticas: {stats_error}")
                        # Criar aba de estatísticas básicas sem dados do MongoDB
                        self._create_basic_statistics_sheet(writer, len(all_publications), len(researchers_summary))
            
            print(f"✅ Excel consolidado exportado: {filename}")
            return filename
            
        except Exception as e:
            print(f"❌ Erro ao exportar Excel consolidado: {e}")
            return None
    
    def _find_keywords_in_publication(self, publication: Dict[str, Any]) -> List[str]:
        """Encontrar keywords relacionadas ao envelhecimento na publicação"""
        keywords_found = []
        
        # Texto para busca (título + resumo/snippet se disponível)
        search_text = ""
        if publication.get("title"):
            search_text += publication["title"].lower()
        if publication.get("snippet"):
            search_text += " " + publication["snippet"].lower()
        
        # Buscar cada keyword
        for keyword in self.KEYWORDS:
            if keyword.lower() in search_text:
                keywords_found.append(keyword)
        
        return list(set(keywords_found))  # Remover duplicatas
    
    def _format_publications_sheet(self, sheet, df):
        """Formatar aba de publicações"""
        # Cabeçalho
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Ajustar larguras das colunas
        column_widths = {
            'A': 25,  # Pesquisador
            'B': 30,  # Instituição
            'C': 50,  # Título
            'D': 30,  # Autores
            'E': 25,  # Publicação
            'F': 8,   # Ano
            'G': 10,  # Citações
            'H': 12,  # Tipo
            'I': 12,  # Plataforma
            'J': 40,  # Keywords
            'K': 12   # Data
        }
        
        for col, width in column_widths.items():
            sheet.column_dimensions[col].width = width
        
        # Bordas
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in sheet.iter_rows(min_row=1, max_row=len(df) + 1):
            for cell in row:
                cell.border = thin_border
    
    def _format_researchers_sheet(self, sheet, df):
        """Formatar aba de pesquisadores"""
        # Cabeçalho
        header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Ajustar larguras
        column_widths = {
            'A': 25,  # Nome
            'B': 35,  # Instituição
            'C': 10,  # H-Index
            'D': 10,  # i10-Index
            'E': 15,  # Citações
            'F': 12,  # Plataforma
            'G': 12,  # Total Pubs
            'H': 12,  # Data
            'I': 12   # Tempo
        }
        
        for col, width in column_widths.items():
            sheet.column_dimensions[col].width = width
    
    def _create_statistics_sheet(self, writer, stats: Dict, total_pubs: int, total_researchers: int):
        """Criar aba de estatísticas"""
        # Dados das estatísticas
        stats_data = [
            ["Métrica", "Valor"],
            ["Total de Pesquisas no Banco", stats.get("total_searches", 0)],
            ["Pesquisas com Filtro de Keywords", stats.get("filtered_searches", 0)],
            ["Total de Publicações no Banco", stats.get("total_publications", 0)],
            ["Publicações neste Export", total_pubs],
            ["Pesquisadores neste Export", total_researchers],
            ["Plataformas Utilizadas", ", ".join(stats.get("platforms", []))],
            ["Última Pesquisa", str(stats.get("latest_search", "N/A")).split("T")[0] if stats.get("latest_search") else "N/A"],
            ["Data deste Export", datetime.now().strftime("%Y-%m-%d %H:%M:%S")]
        ]
        
        df_stats = pd.DataFrame(stats_data[1:], columns=stats_data[0])
        df_stats.to_excel(writer, sheet_name='Estatísticas', index=False)
        
        # Formatar
        sheet = writer.sheets['Estatísticas']
        
        # Cabeçalho
        header_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        header_font = Font(bold=True)
        
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        # Larguras
        sheet.column_dimensions['A'].width = 35
        sheet.column_dimensions['B'].width = 30
    
    def _create_basic_statistics_sheet(self, writer, total_pubs: int, total_researchers: int):
        """Criar aba de estatísticas básicas quando MongoDB não está disponível"""
        # Dados básicos das estatísticas
        stats_data = [
            ["Métrica", "Valor"],
            ["Publicações neste Export", total_pubs],
            ["Pesquisadores neste Export", total_researchers],
            ["Data deste Export", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["Status", "Export gerado com sucesso"],
            ["Observação", "Estatísticas detalhadas indisponíveis"]
        ]
        
        df_stats = pd.DataFrame(stats_data[1:], columns=stats_data[0])
        df_stats.to_excel(writer, sheet_name='Estatísticas', index=False)
        
        # Formatar
        sheet = writer.sheets['Estatísticas']
        
        # Cabeçalho
        header_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        header_font = Font(bold=True)
        
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        # Larguras
        sheet.column_dimensions['A'].width = 35
        sheet.column_dimensions['B'].width = 30

# Instância global
consolidated_exporter = ConsolidatedExcelExporter()